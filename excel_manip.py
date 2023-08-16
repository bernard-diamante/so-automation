import os
import shutil

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

def create_directory(dir_name):
    """
    Creates a directory in the same parent directory with a given name.

    Parameters:
    dir_name (str): The name of the directory to create.
                    Default is "xlsx" for the converted Excel files.
    """
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)
    else:
        shutil.rmtree(dir_name)
        os.makedirs(dir_name)

def create_sheet(output_workbook, sheet_name):
    """
    Creates a sheet in the Excel workbook with a given name

    Parameters:
    outbook_workbook (Workbook): The workbook where the sheet will be created.
    sheet_name (str): The name of the sheet to be created.

    Returns:
    Worksheet: Created sheet
    """
    output_sheet = output_workbook.create_sheet(sheet_name)
    return output_sheet

def convert_xls_to_xlsx(input_folder, output_folder):
    """
    Converts all .xls files in a directory, outputs
    .xlsx to another directory, and returns a list
    of the output file names.

    Parameters:
    input_folder (str): File path of directory containing .xls files.
    output_folder (str): File path of directory where .xlsx files are created.

    Returns:
    list: The list of .xlsx file names created

    Raises:
    ValueError: If there are no service files found in the directory.
    """
    xls_files = os.listdir(input_folder)
    xls_files = [file for file in xls_files
                 if file.startswith("Service_")
                 and file.endswith(".xls")]
    if not xls_files:
        raise ValueError(f"No service files found in the following directory: {input_folder}")


    for xls_file in xls_files:
        input_path = os.path.join(input_folder, xls_file)
        output_path = os.path.join(output_folder, xls_file.replace(".xls", ".xlsx"))

        # Copy over data into the new .xlsx file
        df_xls = pd.read_excel(input_path)

        wb_xlsx = Workbook()
        sheet_xlsx = wb_xlsx.active

        for row in df_xls.iterrows():
            sheet_xlsx.append(row[1].tolist())

        wb_xlsx.save(output_path)

    service_files = [file for file in os.listdir(output_folder)]

    return service_files

def extract_cell(file_path, cell_reference):
    """
    Returns data from a cell in the given file found
    through the input file path.

    Parameters:
    file_path (str): Relative path to the .xlsx input file.
    cell_reference (str): The location of the cell to extract.

    Returns:
    Various: The data from the cell extracted.
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        if cell_reference == None:
            return None
        return sheet[cell_reference].value
    except Exception as e:
        print(f"Error: {e}")

def auto_size_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)  # Add a little extra padding
        column_letter = get_column_letter(cell.column)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def duplicate_excel_file(source_path, destination_path):
    try:
        # Copy the source Excel file to the destination with a new name
        shutil.copy(source_path, destination_path)
        print(f"File duplicated as: {destination_path}")
    except Exception as e:
        print(f"Error: {e}")

def set_list_of_pivot_tables_refresh_on_load(workbook_path):
    workbook = load_workbook(workbook_path)
    for sheet in workbook:
        for pivot in sheet._pivots:
            pivot.cache.refreshOnLoad = True
    workbook.save(workbook_path)

def get_cell_reference(row, column):
    column_letter = get_column_letter(column)
    return f"{column_letter}{row}"

def get_sheet_dimensions(workbook_path, sheet_name):
    workbook = load_workbook(workbook_path)
    sheet = workbook[sheet_name]
    return (sheet.max_row, sheet.max_column)