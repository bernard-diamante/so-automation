import os

from excel_manip import extract_cell, create_table, get_sheet_dimensions, get_cell_reference
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

def populate_raw_data_sheet(file_path, service_files, input_dir):
    """
    Creates and populates the "raw" sheet with data from the service files.
    Population of data is by row/vessel entry.

    Parameters:
    file_path (str): The path of the output file.
    service_files (list): List of service file names; datasource
    raw_cells_to_extract (dict): Map of column headers to extract with cell references of the data.
    input_dir (str): The name of the directory containing the input files


    """

    def find_cell_value_to_right(service_file_path, search_string):
        """
        Finds the specified search string and returns the value to the right.
        """
        try:
            workbook = load_workbook(service_file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                if row[2] == search_string:  # Column C is index 2 (0-based index)
                    # Return value from column D (index 3)
                    return row[3]
            
            # If the search_string is not found, return None
            return None
        except Exception as e:
            print(f"Error: {e}")
            return None
    
    def find_cell_value_to_below(service_file_path, search_string):
        workbook = load_workbook(service_file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=25, min_col=3):
            for cell in row:
                if cell.value == search_string:
                    row_index = cell.row
                    column_index = cell.column
                    value_below = sheet.cell(row=row_index+1, column=column_index).value
                    return value_below

        return None  # Keyword not found or cell below is empty

    def list_vesselnames_cell_references(service_file_path):
        """
        Given an Excel file, return a list of all 
        cell references of the vessel names.

        Parameters:
        service_file_path (str): Relative path to the .xlsx input file.
        """
        workbook = load_workbook(service_file_path)
        sheet = workbook.active
        start_extraction = False
        extracted_cells = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=1, min_col=3, max_col=3, values_only=True), start=1):
            cell_value = row[0]
            if start_extraction:
                if cell_value is None:
                    break
                extracted_cells.append(f"C{row_num}")
            elif cell_value == "Vessel name":
                start_extraction = True

        return extracted_cells

    
    # Field list
    raw_cells_to_extract = {
        "SERVICE NAME": None,
        "SERVICE DESC": "D3",
        "ROUTE": None,
        "LEAD SL": None,
        "SAILING FREQ": None,
        "PARTICIPANTS": None,
        "VESSEL OPERATOR":None,
        "# OF VESSELS": None,
        "# OF VESSELS PER ROW COUNT": None,
        "WEEKLY CAPACITY":None,
        "SHIPS USED": None,
        "PORT ROTATION": None,
        "VESSEL SIZE": None,
        "VESSEL NAME": None
        }
    raw_headers_internal = [
        "PORT",
        "MICT SERVICE NAME",
        "ALT SRVC CD"
        ]
    
    raw_headers_extract = [cell for cell in raw_cells_to_extract.keys()]
    raw_headers_list = list(raw_headers_internal[:2]) + list(raw_headers_extract[:12]) + [raw_headers_internal[2]] + list(raw_headers_extract[12:])
    # Create sheet and populate header
    workbook = load_workbook(file_path)
    raw_data_sheet_name = "raw"

    try:
        sheet_names = workbook.sheetnames
        del workbook[raw_data_sheet_name]
        workbook.create_sheet(raw_data_sheet_name)
    except KeyError:
        # Handle the case where the sheet doesn't exist
        print(f"The sheet '{raw_data_sheet_name}' doesn't exist.")
        return
    finally:
        workbook.save(file_path)
    
    workbook = load_workbook(file_path)
    raw_data_sheet = workbook[raw_data_sheet_name]
    workbook.active = raw_data_sheet
    workbook.title = "Service Overview"
    raw_data_sheet.append(raw_headers_list)
    raw_data_sheet.freeze_panes = "A2"

    def get_column_values(service_file_path, column):
        workbook = load_workbook(service_file_path)
        sheet = workbook.active
        values_set = set()
        
        for cell in sheet[column]:
            if cell.value is None:
                break
            values_set.add(cell.value)

        return values_set

    # Load N4 Services into set
    n4_svcs = get_column_values("n4_svcs.xlsx", "A")

    def find_port(comment_string):
        comment_string = comment_string.lower()
        port_mapping = {
            "north and south": "MICT + ATI",
            "north": "MICT",
            "south": "ATI"
        }

        for search_string, port in port_mapping.items():
            if search_string in comment_string:
                return port

        return "domestic"
        
    def extract_text_between_phrases(input_string, start_phrase, end_phrase):
        """
        Extracts the text between two given phrases in the input string.

        Parameters:
        input_string (str): The input string to search within.
        start_phrase (str): The phrase that marks the start of the desired text.
        end_phrase (str): The phrase that marks the end of the desired text.

        Returns:
        str: The extracted text between the start and end phrases, or an empty string if no text is found.
        """
        start_index = input_string.find(start_phrase) + len(start_phrase)
        end_index = input_string.find(end_phrase)

        return input_string[start_index:end_index].strip() if start_index != -1 and end_index != -1 else ""

    def get_service_name(service_desc):
        if has_parentheses(service_desc):
            cell_value = get_text_in_last_parentheses(service_desc)
            if has_dash(cell_value):
                cell_value = get_text_after_last_dash(cell_value)[:-1]
        else:
            cell_value = get_text_after_last_dash(service_desc)
        return cell_value
        
    def get_mict_service_name(n4_svcs):
        service_name = row_data["SERVICE NAME"]
        port = row_data["PORT"]

        if "MICT" in port:
            if service_name in n4_svcs:
                return service_name
            else:
                return "MANUAL CHECK"
        elif "MICT" not in port:
            return service_name
        
    def strip_lead_sl(service_desc):
        dashes_list = [" - ", " – ", " — "]
        for dash in dashes_list:
            index = service_desc.find(dash)
            if index != -1:
                extract = service_desc[:index]
        delim = " / "
        if delim in extract:
            extract = extract.split(delim)[0].strip()
        return extract

    # Iterate through the .xlsx files and extract cell data
    for service_file in service_files:
        service_file_path = os.path.join(input_dir, service_file)
        row_data = {key: None for key in raw_headers_list}
        
        # PORT
        lookup = "PORT"
        cell_value = find_cell_value_to_below(service_file_path, "Comments")
        start_phrase = "Manila called at"
        end_phrase = "Comments - Service Chronology"
        sliced_cell_value = extract_text_between_phrases(cell_value, start_phrase, end_phrase)
        port = find_port(sliced_cell_value)
        row_data[lookup] = port

        # SERVICE DESC
        lookup = "SERVICE DESC"
        cell_reference = raw_cells_to_extract[lookup]
        cell_value = extract_cell(service_file_path, cell_reference)
        row_data[lookup] = cell_value

        # SERVICE NAME
        def has_parentheses(s):
            return '(' in s and ')' in s
        
        def has_dash(s):
            return '-' in s or "–" in s or "—" in s
        
        def get_text_in_last_parentheses(text):
            last_open = text.rfind("(")
            last_close = text.rfind(")")
            if last_open != -1 and last_close > last_open:
                return text[last_open + 1:last_close]
            else:
                return ""
        
        def get_text_after_last_dash(text):
            last_dash_idx = 0
            idx_temp = 0
            for char in ["-", "–", "—"]:
                temp = service_desc.rfind(char)
                if temp > idx_temp:
                    last_dash_idx = temp
            extracted_text = service_desc[last_dash_idx+1:].strip()
            return extracted_text
            

        # SERVICE NAME
        lookup = "SERVICE NAME"
        service_desc = row_data["SERVICE DESC"]
        cell_value = get_service_name(service_desc)
        row_data[lookup] = cell_value
        print(row_data[lookup])

        # MICT SERVICE NAME
        lookup = "MICT SERVICE NAME"
        cell_value = get_mict_service_name(n4_svcs)
        row_data[lookup] = cell_value

        # ROUTE
        lookup = "ROUTE"
        cell_value = find_cell_value_to_right(service_file_path, "Coverage")
        row_data[lookup] = cell_value

        # LEAD SL
        lookup = "LEAD SL"
        cell_value = row_data["SERVICE DESC"]
        lead_sl = strip_lead_sl(cell_value)
        row_data[lookup] = lead_sl

        # SAILING FREQ
        lookup = "SAILING FREQ"
        cell_value = find_cell_value_to_right(service_file_path, "Sailing frequency")
        row_data[lookup] = cell_value

        def list_participant_by_type(service_file_path, column, type):
            workbook = load_workbook(service_file_path)
            sheet = workbook.active
            cell_values = []

            for cell in sheet[column]:
                if cell.value != None:
                    cell_to_the_right = sheet.cell(row=cell.row, column=cell.column + 1)
                    if cell_to_the_right.value == type:
                        cell_values.append(cell.value)

            return cell_values
        
        
        def format_participants_list(service_file_path, column):
            participant_types = ["Vessel provider", "Slotter"]
            formatted_participants_list = ""
            for participant_type in participant_types:
                participant_list_by_type = list_participant_by_type(service_file_path, column, participant_type)
                if participant_list_by_type:
                    delimited_string = " / ".join(participant_list_by_type)
                    cleaned_string = f"{participant_type}s: {delimited_string}"
                    if participant_type == "Slotter":
                        formatted_participants_list += " / "
                    formatted_participants_list += cleaned_string
            return formatted_participants_list

        # PARTICIPANTS
        lookup = "PARTICIPANTS"
        cell_value = format_participants_list(service_file_path, "C")
        row_data[lookup] = cell_value

        # WEEKLY CAPACITY
        lookup = "WEEKLY CAPACITY"
        cell_value = find_cell_value_to_right(service_file_path, "Weekly capacity (teu)")
        row_data[lookup] = cell_value

        # SHIPS USED
        lookup = "SHIPS USED"
        cell_value = find_cell_value_to_right(service_file_path, "Proforma fleet")
        row_data[lookup] = cell_value

        def extract_vessel_size(input_string):
            # Split the input_string at the word "from"
            parts = input_string.split("from", 1)

            # Check if "from" exists in the input_string
            if len(parts) > 1:
                # Extract everything after "from" and remove leading/trailing spaces
                result = parts[1].strip()
                return result
            else:
                # If "from" is not found, return None
                return None
            
        # VESSEL SIZE
        lookup = "VESSEL SIZE"
        try:
            cell_value = extract_vessel_size(row_data["SHIPS USED"])[:-1]
        except TypeError:
            cell_value = "-"
        row_data[lookup] = cell_value

        # # OF VESSELS
        lookup = "# OF VESSELS"
        try:
            cell_value = int(row_data["SHIPS USED"].split()[0])
        except ValueError:
            cell_value = row_data["SHIPS USED"].split()[0]

        row_data[lookup] = cell_value

        # # OF VESSELS PER ROW COUNT
        lookup = "# OF VESSELS PER ROW COUNT"
        cell_value = 1
        row_data[lookup] = cell_value

        # PORT ROTATION
        lookup = "PORT ROTATION"
        cell_value = find_cell_value_to_below(service_file_path, "Port rotation")
        row_data[lookup] = cell_value

        # WEEKLY CAPACITY
        lookup = "WEEKLY CAPACITY"
        cell_value = find_cell_value_to_right(service_file_path, "Weekly capacity (teu)")
        row_data[lookup] = cell_value

        # VESSEL NAME, VESSEL OPERATOR
        lookup = "VESSEL NAME"
        operator = "VESSEL OPERATOR"
        vessel_name_coordinates_list = list_vesselnames_cell_references(service_file_path)
        # If no vessels listed, default to -
        if not vessel_name_coordinates_list:
            cell_value = "-"
            row_data[lookup] = cell_value
            raw_data_sheet.append([value for value in row_data.values()])
        else:
            # Fill unique fields (VESSEL NAME, VESSEL OPERATOR)
            for cell_reference in vessel_name_coordinates_list:
                cell_value = extract_cell(service_file_path, cell_reference)
                row_data[lookup] = cell_value

                cell_reference = "K" + cell_reference[1:]
                cell_value = extract_cell(service_file_path, cell_reference)
                row_data[operator] = cell_value
            
                # Append row to sheet
                raw_data_sheet.append([value for value in row_data.values()])
    
    # # Set raw data as table
    # sheet_dimensions = get_sheet_dimensions(file_path, raw_data_sheet_name)
    # start_cell = "A1"
    # end_cell = f"{get_cell_reference(sheet_dimensions[0], sheet_dimensions[1])}"
    # cell_range = f"{start_cell}:{end_cell}"
    
    workbook.save(file_path)
    return workbook
